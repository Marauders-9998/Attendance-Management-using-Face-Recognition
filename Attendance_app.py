import os
import sys
import subprocess
from datetime import datetime
import tkinter as tk
from tkinter import messagebox, END
from tkinter import font as tkfont
from calendar import monthrange

import cv2
import pyautogui
from openpyxl import Workbook, load_workbook
import excel
from capture import capture
from capture import detect_faces
from face_train import FaceTrain
from students_list import StudentsList
from main_file import MainFile

class_codes = ['Marauders']
manager_id = 'ADMIN'
manager_pass = 'ubuntu'

current_class = 'Marauders'
if current_class != 'Marauders':
    current_class_obj = MainFile(current_class)
    FaceTrainObj = FaceTrain(current_class)
else:
    current_class_obj = None
    FaceTrainObj = None

# Dynamic variables
normal_width = 1920
normal_height = 1080
x, y = pyautogui.size()

percentage_width = x/(normal_width/100)
percentage_height = y/(normal_height/100)
scale_factor = ((percentage_width + percentage_height)/2)/100

title_fontsize = int(30 * scale_factor)
minimum_size = 24
title_fontsize = max(title_fontsize, minimum_size)

text_fontsize = int(18 * scale_factor)
minimum_textsize = 13
text_fontsize = max(minimum_textsize, text_fontsize)


class SampleApp(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.title_font = tkfont.Font(family='Helvetica', size=18,
                                      weight="bold", slant="italic")
        geo = str(int(0.43 * x)) + 'x' + str(int(0.52 * y))
        self.geometry(geo)
        self.resizable(False, False)
        self.title('Attendance Management App')
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        # The container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.createInitialDirectories()

        self.frames = {}
        for F in (StartPage, StudentPanelPage, ManagerPanelPage,
                  CreateNewBatchPage, AddStudentPage):
            page_name = F.__name__

            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # Put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("StartPage")

    def createInitialDirectories(self):

        temp_dir = os.path.join(os.getcwd(), 'images', '.temp')
        temp_dir_exists = os.path.isdir(temp_dir)

        extras_dir = os.path.join(os.getcwd(), 'extras')
        extras_dir_exists = os.path.isdir(extras_dir)

        studs_list_dir = os.path.join(os.getcwd(), "student's list")
        studs_list_dir_exists = os.path.isdir(studs_list_dir)

        if not extras_dir_exists:
            os.makedirs(extras_dir)
        if not temp_dir_exists:
            os.makedirs(temp_dir)
        if not studs_list_dir_exists:
            os.makedirs(studs_list_dir)

        classes_file = os.path.join(os.getcwd(), 'extras', 'classes.xlsx')
        try:
            global class_codes, current_class
            class_codes = set(['Marauders'])
            wb = load_workbook(classes_file)
            ws = wb.active
            number_of_classes = ws['A1'].value
            if number_of_classes != 0:
                current_class = ws['A2'].value
            for i in range(2, number_of_classes + 2):
                class_codes.add(ws['A' + str(i)].value)
        except:
            wb = Workbook()
            wb.guess_types = True
            ws = wb.active
            ws['A1'] = 0
            wb.save(classes_file)

    def show_frame(self, page_name):
        # Show a frame for the given page name
        frame = self.frames[page_name]
        frame.tkraise()

    def on_closing(self):
        self.destroy()
        self.quit()

    def validate(self, page_name, c_code, id, pwd):
        global class_codes
        if id == manager_id and pwd == manager_pass and c_code in class_codes:
            frame = self.frames[page_name]
            global current_class
            current_class = c_code
            print("current class", current_class)
            if current_class != 'Marauders':
                # frame.lb_class_code.config(text="Batch Code : {}".format(current_class))
                # frame.lb_class_code['text']="Batch Code : {}".format(current_class)
                StudentsList(current_class).make_pkl_file()
                if page_name == "ManagerPanelPage":
                    global FaceTrainObj
                    FaceTrainObj = FaceTrain(current_class)
                elif page_name == "StudentPanelPage":
                    global current_class_obj
                    current_class_obj = MainFile(current_class)
            elif current_class == 'Marauders' and page_name == 'StudentPanelPage':
                messagebox.showerror('Error', 'Only Manager Panel is valid for this class-code')
                return
            else:
                frame = self.frames['CreateNewBatchPage']
            frame.tkraise()
        else:
            messagebox.showerror('Error', 'Incorrect Credentials !!')


class StartPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.bkg = '#222629'
        self.text_color = '#65CCB8'
        StartPage.config(self, bg=self.bkg)

        entry_width = int(30 * scale_factor)
        min_entrywidth = 25
        entry_width = max(entry_width, min_entrywidth)
        self.label = tk.Label(self, width=25, text="Attendance Management",
                              bg=self.bkg, fg="#F8E9A1",
                              font=("Times", title_fontsize))
        self.label.place(x=130 * scale_factor, y=40 * scale_factor)

        self.lb_class = tk.Label(self, text="CLASS-CODE: ", bg=self.bkg,
                                 fg=self.text_color,
                                 font=("Courier", text_fontsize))
        self.lb_class.place(x=170 * scale_factor, y=155 * scale_factor)
        self.tv_class = tk.Entry(self, width=entry_width)
        self.tv_class.focus()
        self.tv_class.place(x=340 * scale_factor, y=160 * scale_factor)

        self.lb_username = tk.Label(self, text="USERNAME : ", bg=self.bkg,
                                    fg=self.text_color,
                                    font=("Courier", text_fontsize))
        self.lb_username.place(x=183 * scale_factor, y=210 * scale_factor)
        self.tv_username = tk.Entry(self, width=entry_width)
        self.tv_username.place(x=340 * scale_factor, y=215 * scale_factor)

        self.lb_pass = tk.Label(self, text="PASSWORD : ", bg=self.bkg,
                                fg=self.text_color,
                                font=("Courier", text_fontsize))
        self.lb_pass.place(x=183 * scale_factor, y=265 * scale_factor)
        self.tv_pass = tk.Entry(self, show="*", width=entry_width)
        self.tv_pass.place(x=340 * scale_factor, y=270 * scale_factor)
        self.showbtn = tk.Button(self, text="SHOW", bg="#ed3833",
                                 command=self.show, width=2,
                                 font=("", 8))  # 32ff6a
        self.show = False
        self.showbtn.place(x=(340 * scale_factor) + (entry_width * 9),
                           y=270*scale_factor)

        bt_width = int(16 * scale_factor)
        min_width = 12
        bt_width = max(min_width, bt_width)
        self.button1 = tk.Button(
            self, bg="#45056e", fg=self.text_color,
            text="Go to\nStudent Portal", width=bt_width, height=3,
            font=("", 15),
            command=lambda: self.doWork(
                "StudentPanelPage",
                self.tv_class.get(),
                self.tv_username.get(),
                self.tv_pass.get()
            ))
        self.button2 = tk.Button(
            self, bg="#5f1854", fg=self.text_color,
            text="Go to\nManager Portal",
            width=bt_width, height=3, font=("", 15),
            command=lambda: self.doWork(
                "ManagerPanelPage",
                self.tv_class.get(),
                self.tv_username.get(),
                self.tv_pass.get()
            ))
        self.button1.place(x=140 * scale_factor, y=340 * scale_factor)
        self.button2.place(x=410 * scale_factor, y=340 * scale_factor)
        self.bt_exit = tk.Button(self, bg="red", fg="yellow", text="Exit",
                                 width=10, command=self.exit)
        self.bt_exit.place(x=340 * scale_factor, y=490 * scale_factor)

    def show(self):
        if self.show is False:
            self.tv_pass.config(show="")
            self.show = True
            self.showbtn.config(bg="#32ff6a")
        elif self.show is True:
            self.tv_pass.config(show="*")
            self.show = False
            self.showbtn.config(bg="#ed3833")

    def doWork(self, page_name, c_code, id, pwd):
        self.tv_class.delete(0, END)
        self.tv_username.delete(0, END)
        self.tv_pass.delete(0, END)
        self.controller.validate(page_name, c_code, id, pwd)

    def exit(self):
        self.controller.destroy()
        self.controller.quit()


class StudentPanelPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.bkg = '#222629'
        self.text_color = '#65CCB8'
        StudentPanelPage.config(self, bg=self.bkg)

        label = tk.Label(self, width=25, text="Student Portal", bg=self.bkg,
                         fg="#F8E9A1", font=("Times", title_fontsize))
        label.place(x=130 * scale_factor, y=40 * scale_factor)
        bt_mark = tk.Button(self, bg="#45056e", fg=self.text_color,
                            text="Mark my\n ATTENDANCE", font=("", 18),
                            width=16, height=7, command=self.doWork)
        bt_mark.place(x=215 * scale_factor, y=150 * scale_factor)

        bt_back = tk.Button(self, bg="#af0404", fg="yellow", text="Back",
                            width=10, command=lambda: controller.show_frame(
                                "StartPage"))
        bt_back.place(x=115 * scale_factor, y=485 * scale_factor)
        bt_exit = tk.Button(self, bg="#af0404", fg="yellow", text="Exit",
                            width=10, command=self.exit)
        bt_exit.place(x=520 * scale_factor, y=485 * scale_factor)

    def doWork(self):
        global current_class_obj
        current_class_obj.capture_and_mark()

    def exit(self):
        self.controller.destroy()
        self.controller.quit()


class ManagerPanelPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.bkg = '#222629'
        self.text_color = '#65CCB8'
        ManagerPanelPage.config(self, bg=self.bkg)
        global current_class

        label = tk.Label(self, width=25, text="Manager Portal", bg=self.bkg,
                         fg="#F8E9A1", font=("Times", title_fontsize))
        label.place(x=130 * scale_factor, y=40 * scale_factor)

        bt_ht = int(5 * scale_factor)
        min_ht = 3
        bt_ht = max(min_ht, bt_ht)
        bt_width = int(15 * scale_factor)
        min_width = 12
        bt_width = max(min_width, bt_width)
        bt_train = tk.Button(self, text="Train the\n Recogniser",
                             bg="#45056e", fg=self.text_color, width=bt_width,
                             height=bt_ht, font=("", 15), command=self.doWork)
        bt_train.place(x=130 * scale_factor, y=160 * scale_factor)
        bt_addstud = tk.Button(self, text="Add a student\n to this class",
                               bg="#3b0944", width=bt_width,
                               fg=self.text_color, height=bt_ht,
                               font=("", 15), command=self.addStudent)
        bt_addstud.place(x=430 * scale_factor, y=160 * scale_factor)
        bt_view_register = tk.Button(self, text="View\nAttendance\nRegister",
                                     width=bt_width, bg="#581845",
                                     fg=self.text_color, height=bt_ht,
                                     font=("", 15), command=self.viewRegister)
        bt_view_register.place(x=265 * scale_factor, y=320 * scale_factor)

        bt_back = tk.Button(self, bg="#af0404", fg="yellow", text="Back",
                            width=10, command=lambda: controller.show_frame(
                                "StartPage"))
        bt_back.place(x=115 * scale_factor, y=485 * scale_factor)
        bt_exit = tk.Button(self, bg="#af0404", fg="yellow", text="Exit",
                            width=10, command=self.exit)
        bt_exit.place(x=520 * scale_factor, y=485 * scale_factor)

    def viewRegister(self):
        global current_class
        atten_register = os.path.join(os.getcwd(), 'extras', current_class, f'{current_class}.xlsx')
        try:
            opener = 'open' if sys.platform == 'darwin' else 'xdg-open'
            subprocess.call([opener, atten_register])
        except (subprocess.CalledProcessError, OSError):
            os.startfile(atten_register)

    def addStudent(self):
        global current_class
        students_list_file = os.path.join(os.getcwd(), "student's list", f'{current_class}.xlsx')
        self.controller.show_frame("AddStudentPage")

    def doWork(self):
        global FaceTrainObj, current_class
        if current_class == 'Marauders':
            wb = load_workbook(os.path.join(os.getcwd(), 'extras', 'classes.xlsx'))
            ws = wb.active
            number_of_classes = ws['A1'].value
            if number_of_classes == 0:
                messagebox.showerror('Error', 'No class in the database yet')
                return
            else:
                messagebox.showerror('Error', 'Please login again with a valid class-code')
                current_class = ws['A2'].value
                return

        FaceTrainObj.main()

    def exit(self):
        self.controller.destroy()
        self.controller.quit()


class CreateNewBatchPage(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.bkg = '#222629'
        self.text_color = '#65CCB8'
        CreateNewBatchPage.config(self, bg=self.bkg)

        self.label = tk.Label(self, width=25, text="Create a new Batch",
                              bg=self.bkg, fg="#F8E9A1",
                              font=("Times", title_fontsize))
        self.label.place(x=130 * scale_factor, y=40 * scale_factor)

        entry_width = int(30 * scale_factor)
        min_entrywidth = 23
        entry_width = max(min_entrywidth, entry_width)
        self.lb_class = tk.Label(self, text="CLASS-CODE: ", bg=self.bkg,
                                 fg=self.text_color,
                                 font=("Courier", text_fontsize))
        self.lb_class.place(x=320 * scale_factor, y=140 * scale_factor)
        self.tv_class = tk.Entry(self, width=entry_width, font=("", 16),
                                 justify='center')
        self.tv_class.focus()
        self.tv_class.place(x=190 * scale_factor, y=175 * scale_factor)

        self.lb_number = tk.Label(self, text="NUMBER OF STUDENTS: ",
                                  bg=self.bkg, fg=self.text_color,
                                  font=("Courier", text_fontsize))
        self.lb_number.place(x=265 * scale_factor, y=240 * scale_factor)
        vcmd = (self.controller.register(self.validate_number_field),
                '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')
        self.tv_number = tk.Entry(self, width=entry_width, font=("", 16),
                                  justify='center', validate='key',
                                  validatecommand=vcmd)
        self.tv_number.place(x=190 * scale_factor, y=275 * scale_factor)

        bt_new_batch = tk.Button(self, bg="#45056e", fg=self.text_color,
                                 text="ADD BATCH", width=16, height=2,
                                 font=("", 15), command=self.create_new_batch)
        bt_new_batch.place(x=237 * scale_factor, y=350 * scale_factor)

        bt_back = tk.Button(self, bg="#ed3833", fg="yellow", text="Back",
                            width=10, command=lambda: controller.show_frame(
                                "StartPage"))
        bt_back.place(x=340 * scale_factor, y=490 * scale_factor)

    def validate_number_field(self, action, index, value_if_allowed,
                              prior_value, text, validation_type,
                              trigger_type, widget_name):
        if text in '0123456789':
            try:
                int(value_if_allowed)
                return True
            except (TypeError, ValueError):
                return False
        else:
            return False

    def create_new_batch(self):
        global class_codes
        class_name = self.tv_class.get()
        number_of_studs = int(self.tv_number.get())

        batchexists = os.path.exists(os.path.join(os.getcwd(), 'extras', class_name))
        if batchexists:
            messagebox.showerror('Error', 'This batch name already exists')
            return

        if number_of_studs < 1 or number_of_studs > 99:
            messagebox.showerror('Error', "Number of students not in allowed range!")
            return

        images_dir = os.path.join(os.getcwd(), 'images', class_name)
        unrecog_studs_dir = os.path.join(os.getcwd(), 'images', class_name,
                                         'unrecognized students')
        os.makedirs(unrecog_studs_dir)

        for i in range(number_of_studs):
            os.makedirs(os.path.join(images_dir, 's' + str(i).zfill(2)))

        studs_list_file = os.path.join(os.getcwd(), "student's list", f'{class_name}.xlsx')
        wb = Workbook()
        wb.guess_types = True
        ws = wb.active
        ws['A1'] = 0
        ws['B1'] = number_of_studs
        wb.save(studs_list_file)

        classes_file = os.path.join(os.getcwd(), 'extras', 'classes.xlsx')
        wb = load_workbook(classes_file)
        ws = wb.active
        row = ws['A1'].value
        ws['A1'] = row + 1
        ws['A' + str(row + 2)] = class_name
        wb.save(classes_file)

        sl = StudentsList(class_name)
        sl.make_pkl_file()

        atten_reg_dir = os.path.join(os.getcwd(), 'extras', class_name)
        os.makedirs(atten_reg_dir)
        wb = excel.attendance_workbook(class_name)

        messagebox.showinfo("Batch Successfully Created",
                            "All the necessary Directories and Files created"
                            f"for Batch-Code {class_name}")
        class_codes.add(class_name)
        self.tv_class.delete(0, END)
        self.tv_number.delete(0, END)
        self.tv_number.insert(0, "")
        self.controller.show_frame('StartPage')


class AddStudentPage(tk.Frame):
    """docstring for AddStudentPage"""
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller
        self.bkg = '#222629'
        self.text_color = '#65CCB8'
        AddStudentPage.config(self, bg=self.bkg)

        self.label = tk.Label(self, width=25, text="Add a new Student",
                              bg=self.bkg, fg="#F8E9A1",
                              font=("Times", title_fontsize))
        self.label.place(x=130 * scale_factor, y=40 * scale_factor)

        self.lb_name = tk.Label(self, text="Name of Student: ",
                                bg=self.bkg, fg=self.text_color,
                                font=("Courier", text_fontsize))
        self.lb_name.place(x=285 * scale_factor, y=175 * scale_factor)
        entry_width = int(30 * scale_factor)
        min_entrywidth = 23
        entry_width = max(min_entrywidth, entry_width)
        self.tv_name = tk.Entry(self, width=entry_width, justify='center',
                                font=("", 16))
        self.tv_name.focus()
        self.tv_name.place(x=180 * scale_factor, y=210 * scale_factor)

        self.bt_addstud = tk.Button(self, bg="#45056e", fg=self.text_color,
                                    text="ADD STUDENT", width=16, height=2,
                                    font=("", 15), command=self.doWork)
        self.bt_addstud.place(x=245 * scale_factor, y=300 * scale_factor)

        self.bt_back = tk.Button(self, bg="#af0404", fg="yellow", text="Back",
                                 width=10,
                                 command=lambda: controller.show_frame(
                                     "ManagerPanelPage"))
        self.bt_back.place(x=115 * scale_factor, y=485 * scale_factor)
        self.bt_exit = tk.Button(self, bg="#af0404", fg="yellow",
                                 text="Exit", width=10, command=self.exit)
        self.bt_exit.place(x=520 * scale_factor, y=485 * scale_factor)

    def doWork(self):
        global current_class

        students_list_file = os.path.join(os.getcwd(), "student's list", f'{current_class}.xlsx')
        wb = load_workbook(students_list_file)
        ws = wb.active
        total_studs = ws['B1'].value
        currently_studs = ws['A1'].value
        if currently_studs == total_studs:
            messagebox.showinfo("Batch Full", "No more accomodation for any new student")
            return
        ws['A1'] = currently_studs + 1
        row_here = currently_studs + 2
        student_name = self.tv_name.get()
        ws['A' + str(row_here)] = student_name
        roll_no = currently_studs
        roll_no = str(roll_no).zfill(2)
        ws['B' + str(row_here)] = current_class + roll_no
        wb.save(students_list_file)

        sl = StudentsList(current_class)
        sl.make_pkl_file()

        atten_reg = os.path.join(os.getcwd(), 'extras', current_class, f'{current_class}.xlsx')
        wb = load_workbook(atten_reg)
        today = datetime.now().date()
        month = today.strftime('%B')
        try:
            # If current month exists
            ws = wb[month]
            row_here = excel.gap + int(roll_no)
            ws.merge_cells(start_row=row_here, start_column=3,
                           end_row=row_here, end_column=4)
            ws.cell(row=row_here, column=1, value=str(int(roll_no) + 1))
            fill_roll = excel.PatternFill(fgColor='A5EFAF', fill_type='solid')
            fill_name = excel.PatternFill(fgColor='E9EFA5', fill_type='solid')
            ws.cell(row=row_here, column=2,
                    value=current_class + roll_no).fill = fill_roll
            ws.cell(row=row_here, column=3,
                    value=student_name).fill = fill_name
            cellrange = 'A' + str(row_here) + ':AN' + str(row_here)
            excel.set_border(ws, cellrange)
            first_date_column = excel.get_column_letter(excel.date_column)
            days_in_month = monthrange(today.year, today.month)[1]
            last_date_column = excel.get_column_letter(excel.date_column + days_in_month - 1)
            sum_cell_column = excel.date_column + days_in_month + 1
            fill_total = excel.PatternFill(fgColor='25F741', fill_type='solid')
            ws.cell(row=row_here, column=sum_cell_column,
                    value="=SUM(" + first_date_column + str(row_here) + ":" +
                    last_date_column + str(row_here) + ")").fill = fill_total
            wb.save(atten_reg)
        except:  # If current month does not exist
            wb = excel.attendance_workbook(current_class)

        image_dir = os.path.join(os.getcwd(), 'images', current_class, 's' + roll_no)
        messagebox.showinfo('Student Added',
                            f"{student_name} admitted!\n"
                            "Click OK to proceed for capturing training images. "
                            "Make sure that your surroundings are well lit")
        i = 5
        face_detected_right = False
        xml_file = os.path.join(os.getcwd(), 'haarcascade_frontalface_default.xml')
        while i > 0:
            while face_detected_right is False:
                img_path, frame = capture()
                face_detected_right = detect_faces(xml_file, frame)
            img_path = os.path.join(
                os.getcwd(), 'images', current_class,
                "s"+str(roll_no).zfill(2),
                os.path.basename(img_path)
            )
            cv2.imwrite(img_path, frame)
            cv2.destroyAllWindows()
            face_detected_right = False
            i = i - 1
        if i == 0:
            messagebox.showinfo("Message", "Training images added successfully!")
        messagebox.showinfo('Enrollment Number',
                            f'Roll Number of student: {current_class + roll_no}')
        self.tv_name.delete(0, END)

    def exit(self):
        self.controller.destroy()
        self.controller.quit()


if __name__ == "__main__":
    app = SampleApp()
    app.mainloop()
