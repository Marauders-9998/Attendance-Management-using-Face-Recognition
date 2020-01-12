from os import getcwd
from os.path import join
from calendar import monthrange
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from students_list import StudentsList

HEADING_FONT = Font(size=20, bold=True)
ALIGNMENT = Alignment(horizontal='center', vertical='center')
FILL_DATE = PatternFill(fgColor='AFA5EF', fill_type='solid')
FILL_ROLL = PatternFill(fgColor='A5EFAF', fill_type='solid')
FILL_NAME = PatternFill(fgColor='E9EFA5', fill_type='solid')
FILL_TOTAL = PatternFill(fgColor='25F741', fill_type='solid')
FILL_ABSENT = PatternFill(fgColor='FF1010', fill_type='solid')
FILL_PRESENT = PatternFill(fgColor='10FF10', fill_type='solid')

GAP = 5
DATE_COLUMN = 5

NAMES = ['Marauders']
ROLLS = ['000001']

NUMBER_OF_STUDENTS = len(NAMES)


def todays_date():
    return datetime.now().date()


def month():
    return todays_date().strftime('%B')


def make_file_name(class_name):
    return join(getcwd(), 'extras', class_name, class_name + '.xlsx')


def current_month_exists(wb):
    try:
        ws = wb[month()]
        return True
    except:
        return False


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)
    max_y = len(rows) - 1
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def createWorksheet(wb):
    # "creating register"
    global NAMES, ROLLS, NUMBER_OF_STUDENTS
    manth = month()
    ws = wb.create_sheet(manth)
    ws.merge_cells('A1:AN2')
    heading = 'Attendance Record for ' + manth
    heading = heading + ' ' * 50 + heading + ' ' * 50 + heading
    ws['A1'] = heading
    ws['A1'].font = HEADING_FONT
    ws['A1'].alignment = ALIGNMENT

    # ws.sheet_properties.tabColor = 'FFFF66'
    ws['A3'] = 'S.No'
    ws['B3'] = 'Enrollment'
    ws['B3'].fill = FILL_ROLL
    ws.merge_cells('C3:D3')
    ws['C3'] = 'Name'
    ws['C3'].fill = FILL_NAME

    cellrange = 'A' + str(GAP - 1) + ':AN' + str(GAP - 1)
    set_border(ws, cellrange)
    for i in range(NUMBER_OF_STUDENTS):
        ws.merge_cells(start_row=GAP + i, start_column=3, end_row=GAP + i, end_column=4)
        ws.cell(row=GAP + i, column=1, value=str(i + 1))
        ws.cell(row=GAP + i, column=2, value=ROLLS[i]).fill = FILL_ROLL
        ws.cell(row=GAP + i, column=3, value=NAMES[i]).fill = FILL_NAME

        cellrange = 'A' + str(GAP + i) + ':AN' + str(GAP + i)
        set_border(ws, cellrange)
    today = todays_date()
    first_day = date(today.year, today.month, 1)
    days_in_month = monthrange(today.year, today.month)[1]
    for i in range(days_in_month):
        dt = first_day + timedelta(days=i)
        ws.cell(row=3, column=DATE_COLUMN + i, value=dt).fill = FILL_DATE

    first_date_column = get_column_letter(DATE_COLUMN)
    last_date_column = get_column_letter(DATE_COLUMN + days_in_month - 1)
    sum_cell_column = DATE_COLUMN + days_in_month + 1
    ws.cell(row=3, column=sum_cell_column, value="Total").fill = FILL_TOTAL
    for i in range(NUMBER_OF_STUDENTS):
        ws.cell(row=GAP + i, column=sum_cell_column,
                value="=SUM(" + first_date_column + str(GAP + i) + ":" + last_date_column + str(
                    GAP + i) + ")").fill = FILL_TOTAL


def mark_absent(wb, studs_absent, class_name):
    ws = wb[month()]
    col = todays_date().day + DATE_COLUMN - 1
    for i in range(NUMBER_OF_STUDENTS):
        if ws.cell(row=GAP + i, column=2).value in studs_absent:
            ws.cell(row=GAP + i, column=col, value=0).fill = FILL_ABSENT
    wb.save(make_file_name(class_name))


def mark_present(wb, studs_present, class_name):
    ws = wb[month()]
    col = todays_date().day + DATE_COLUMN - 1
    for i in range(NUMBER_OF_STUDENTS):
        if ws.cell(row=GAP + i, column=2).value in studs_present:
            ws.cell(row=GAP + i, column=col, value=1).fill = FILL_PRESENT
    wb.save(make_file_name(class_name))


def attendance_workbook(class_name):
    sl = StudentsList(class_name)
    tupl = sl.load_pkl_file()
    global NAMES, ROLLS, NUMBER_OF_STUDENTS
    NAMES = tupl[0]
    ROLLS = tupl[1]
    NUMBER_OF_STUDENTS = len(NAMES)
    # Current length of NAMES is len(NAMES)
    filename = make_file_name(class_name)
    try:
        wb = load_workbook(filename)
    except:
        wb = Workbook()
    wb.guess_types = True
    if not current_month_exists(wb):
        createWorksheet(wb)
    wb.save(filename)
    return wb


if __name__ == '__main__':
    pass
